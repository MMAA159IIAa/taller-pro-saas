import openpyxl
import os

ruta = r"c:\Users\leech\OneDrive\Documentos\RELACION DE CARROS 2026.xlsx"
if os.path.exists(ruta):
    wb = openpyxl.load_workbook(ruta, read_only=True)
    print(f"Hojas encontradas: {wb.sheetnames}")
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        print(f"\n--- Datos de {sheet} (primeras 5 filas) ---")
        for row in ws.iter_rows(max_row=5, values_only=True):
            print(row)
else:
    print("Archivo no encontrado")
