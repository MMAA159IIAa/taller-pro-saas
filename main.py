import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import sqlite3
from datetime import datetime
from tkcalendar import DateEntry
from dateutil.relativedelta import relativedelta
import csv
import os
import sys
import smtplib
import requests
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.chart import BarChart, Reference

# URL de tu servidor central SaaS
API_BASE_URL = "http://localhost:8000"

# ── Agentes y servicios ───────────────────────────────────────────────────────
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
try:
    from agentes.notification_agent import NotificationAgent
    from agentes.sales_agent import SalesAgent
    from agentes.finance_import_agent import FinanceImportAgent
    from agentes.sales_chat_agent import SalesChatAgent
    from agentes.lead_detective_agent import LeadDetectiveAgent
    from utils.database import crear_tablas as _crear_tablas_ext, get_config, set_config
    AGENTES_OK = True
except ImportError as _e:
    AGENTES_OK = False
    print(f"[WARN] Agentes no cargados: {_e}")

BG_DARK      = "#0d0d0d"
BG_CARD      = "#1a1a1a"
BG_INPUT     = "#242424"
ACCENT       = "#e8a020"
ACCENT_DARK  = "#b87d10"
TEXT_PRIMARY = "#f0f0f0"
TEXT_MUTED   = "#888888"
SUCCESS      = "#2ecc71"
DANGER       = "#e74c3c"
ERROR        = "#e74c3c"
BORDER       = "#2a2a2a"
FONT_TITLE   = ("Segoe UI", 22, "bold")
FONT_HEADER  = ("Segoe UI", 11, "bold")
FONT_BODY    = ("Segoe UI", 10)
FONT_SMALL   = ("Segoe UI", 9)

def obtener_ruta_db():
    if getattr(sys, 'frozen', False):
        base_path = os.path.join(os.environ['LOCALAPPDATA'], "TallerPro")
    else:
        base_path = os.path.dirname(os.path.abspath(__file__))
    if not os.path.exists(base_path):
        os.makedirs(base_path)
    return os.path.join(base_path, "taller_pro.db")

DB_NAME = obtener_ruta_db()

def conectar():
    return sqlite3.connect(DB_NAME)

def crear_tablas():
    conn = conectar()
    c = conn.cursor()
    c.execute("""CREATE TABLE IF NOT EXISTS registros (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        nombre TEXT, telefono TEXT, correo TEXT,
        auto TEXT, numero_economico TEXT, placas TEXT,
        servicio TEXT, fecha TEXT, costo REAL DEFAULT 0,
        estatus TEXT DEFAULT 'Pendiente')""")
    c.execute("""CREATE TABLE IF NOT EXISTS finanzas (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        tipo TEXT, concepto TEXT, monto REAL, fecha TEXT, notas TEXT)""")
    c.execute("""CREATE TABLE IF NOT EXISTS config (
        clave TEXT PRIMARY KEY, valor TEXT)""")
    conn.commit()
    conn.close()

def boton(parent, texto, comando, color=ACCENT, texto_color=BG_DARK, w=14):
    return tk.Button(parent, text=texto, command=comando,
                     bg=color, fg=texto_color, font=FONT_HEADER,
                     bd=0, relief="flat", padx=12, pady=7,
                     cursor="hand2", width=w,
                     activebackground=ACCENT_DARK, activeforeground=BG_DARK)

def entrada(parent, width=30):
    e = tk.Entry(parent, bg=BG_INPUT, fg=TEXT_PRIMARY,
                 font=FONT_BODY, bd=0, relief="flat",
                 insertbackground=ACCENT, width=width)
    e.config(highlightthickness=1, highlightbackground=BORDER, highlightcolor=ACCENT)
    return e

class TallerPro(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("TALLER PRO - Sistema de Gestion")
        self.geometry("1400x750")
        self.configure(bg=BG_DARK)
        self.resizable(True, True)
        try:
            ico = os.path.join(os.path.dirname(os.path.abspath(__file__)), "icono.ico")
            if os.path.exists(ico):
                self.iconbitmap(ico)
        except: pass
        crear_tablas()
        if AGENTES_OK:
            _crear_tablas_ext()
        
        # El check de activación ahora está en el bloque __main__ para mayor seguridad
        self.taller_id = get_config("taller_id", "")
        self._validar_licencia()
        
        self._build_ui()
        self.actualizar_dashboard()
        if AGENTES_OK:
            self._iniciar_agentes()

    def _validar_licencia(self):
        try:
            resp = requests.get(f"{API_BASE_URL}/api/v1/license/check/{self.taller_id}", timeout=5).json()
            if resp.get("status") == "suspendido":
                messagebox.showerror("Acceso Denegado", resp.get("mensaje"))
                sys.exit()
            if resp.get("fecha_vencimiento"):
                set_config("fecha_vencimiento", resp["fecha_vencimiento"])
            
            self.status_licencia = resp.get("status", "activo")
        except:
            # Modo Offline
            self.status_licencia = "activo"

    def _iniciar_agentes(self):
        self.notif_agent   = NotificationAgent(intervalo_segundos=30)
        self.sales_agent   = SalesAgent(intervalo_horas=24)
        self.finance_agent = FinanceImportAgent()
        self.chat_agent    = SalesChatAgent()
        self.lead_agent    = LeadDetectiveAgent()
        self.notif_agent.iniciar()
        self.sales_agent.iniciar()
        self.lead_agent.iniciar()

    def _build_ui(self):
        self.sidebar = tk.Frame(self, bg=BG_CARD, width=200)
        self.sidebar.pack(side="left", fill="y")
        self.sidebar.pack_propagate(False)
        self.main = tk.Frame(self, bg=BG_DARK)
        self.main.pack(side="left", fill="both", expand=True)
        self._build_sidebar()
        self.frames = {}
        frame_classes = [FrameDashboard, FrameClientes, FrameFinanzas, FrameReportes, FrameConfig]
        if AGENTES_OK:
            frame_classes += [FrameAgentes, FrameChatbot, FrameLogs]
        for Cls in frame_classes:
            f = Cls(self.main, self)
            self.frames[Cls.__name__] = f
            f.place(relx=0, rely=0, relwidth=1, relheight=1)
        self.mostrar("FrameDashboard")

    def _build_sidebar(self):
        logo_frame = tk.Frame(self.sidebar, bg=BG_CARD, pady=20)
        logo_frame.pack(fill="x")
        tk.Label(logo_frame, text="TALLER PRO", bg=BG_CARD, fg=ACCENT,
                 font=("Segoe UI", 13, "bold")).pack()
        tk.Label(logo_frame, text="Sistema de Gestion", bg=BG_CARD,
                 fg=TEXT_MUTED, font=FONT_SMALL).pack()
        tk.Frame(self.sidebar, bg=BORDER, height=1).pack(fill="x", padx=15)
        nav_items = [
            ("Dashboard",    "FrameDashboard"),
            ("Clientes",     "FrameClientes"),
            ("Finanzas",     "FrameFinanzas"),
            ("Reportes",     "FrameReportes"),
            ("Configuracion","FrameConfig"),
        ]
        if AGENTES_OK:
            nav_items += [
                ("Agentes IA",   "FrameAgentes"),
                ("Chatbot Ventas","FrameChatbot"),
                ("Logs Sistema",  "FrameLogs"),
            ]
        self.nav_buttons = {}
        for texto, frame in nav_items:
            btn = tk.Button(self.sidebar, text=texto, bg=BG_CARD,
                            fg=TEXT_PRIMARY, font=FONT_BODY, bd=0,
                            relief="flat", anchor="w", padx=20, pady=12,
                            cursor="hand2", width=22,
                            activebackground=ACCENT, activeforeground=BG_DARK,
                            command=lambda f=frame: self.mostrar(f))
            btn.pack(fill="x")
            self.nav_buttons[frame] = btn
        tk.Label(self.sidebar, text="v2.0 Pro", bg=BG_CARD,
                 fg=TEXT_MUTED, font=FONT_SMALL).pack(side="bottom", pady=10)

    def mostrar(self, nombre):
        for n, btn in self.nav_buttons.items():
            btn.config(bg=ACCENT if n == nombre else BG_CARD,
                       fg=BG_DARK if n == nombre else TEXT_PRIMARY)
        self.frames[nombre].tkraise()
        if hasattr(self.frames[nombre], "on_show"):
            self.frames[nombre].on_show()

    def actualizar_dashboard(self):
        if "FrameDashboard" in self.frames:
            self.frames["FrameDashboard"].actualizar()


class FrameDashboard(tk.Frame):
    def __init__(self, parent, app):
        super().__init__(parent, bg=BG_DARK)
        self.app = app
        self.kpis = {}
        self._build()
        self.after(1000, self._check_timer_vencimiento) # Iniciar cronómetro tras 1s

    def _build(self):
        # 1. Cabecera con Título y Banner de Vencimiento
        top = tk.Frame(self, bg=BG_DARK, pady=18, padx=25)
        top.pack(fill="x")
        tk.Label(top, text="Dashboard", bg=BG_DARK, fg=TEXT_PRIMARY,
                 font=FONT_TITLE).pack(side="left")
        
        self.lbl_venc = tk.Label(top, text="Cargando suscripción...", bg=BG_DARK, 
                                 fg=TEXT_MUTED, font=("Segoe UI", 9, "bold"))
        self.lbl_venc.pack(side="right", padx=10)

        # 2. Contenedor de KPIs (Tarjetas)
        self.kpi_frame = tk.Frame(self, bg=BG_DARK, padx=25)
        self.kpi_frame.pack(fill="x")
        
        kpi_defs = [
            ("clientes", "Clientes Totales", TEXT_PRIMARY),
            ("hoy",      "Servicios Hoy",    ACCENT),
            ("ingresos", "Ingresos del Mes", SUCCESS),
            ("egresos",  "Egresos del Mes",  DANGER),
            ("balance",  "Balance del Mes",  ACCENT),
            ("vencidos", "Serv. Vencidos",   DANGER),
        ]
        for key, titulo, color in kpi_defs:
            card = tk.Frame(self.kpi_frame, bg=BG_CARD, bd=0,
                            highlightthickness=1, highlightbackground=BORDER)
            card.pack(side="left", expand=True, fill="both", padx=6, pady=4)
            lbl_val = tk.Label(card, text="0", bg=BG_CARD, fg=color,
                               font=("Segoe UI", 20, "bold"))
            lbl_val.pack(pady=(14,2))
            tk.Label(card, text=titulo, bg=BG_CARD, fg=TEXT_MUTED,
                     font=FONT_SMALL).pack(pady=(2,14))
            self.kpis[key] = lbl_val

        # 3. Tabla de Últimos Servicios
        mid = tk.Frame(self, bg=BG_DARK, padx=25, pady=10)
        mid.pack(fill="both", expand=True)
        tk.Label(mid, text="Últimos Servicios", bg=BG_DARK,
                 fg=TEXT_PRIMARY, font=FONT_HEADER).pack(anchor="w")
        
        cols = ("Cliente","Auto","Servicio","Fecha","Costo","Estatus")
        self.tabla = ttk.Treeview(mid, columns=cols, show="headings", height=10)
        
        style = ttk.Style()
        style.theme_use("clam")
        style.configure("Treeview", background=BG_CARD, foreground=TEXT_PRIMARY,
                        fieldbackground=BG_CARD, rowheight=28, font=FONT_BODY)
        style.configure("Treeview.Heading", background=BG_INPUT,
                        foreground=ACCENT, font=FONT_HEADER, relief="flat")
        style.map("Treeview", background=[("selected", ACCENT)],
                  foreground=[("selected", BG_DARK)])
        
        for col in cols:
            self.tabla.heading(col, text=col)
            self.tabla.column(col, width=160)
        self.tabla.pack(fill="both", expand=True, pady=5)

    def on_show(self):
        self.actualizar()

    def actualizar(self):
        try:
            conn = conectar()
            mes = datetime.now().strftime("%Y-%m")
            hoy = datetime.now().strftime("%Y-%m-%d")
            
            total_clientes = conn.execute("SELECT COUNT(*) FROM registros").fetchone()[0]
            servicios_hoy  = conn.execute("SELECT COUNT(*) FROM registros WHERE fecha=?", (hoy,)).fetchone()[0]
            ingresos = conn.execute("SELECT COALESCE(SUM(monto),0) FROM finanzas WHERE tipo='Ingreso' AND fecha LIKE ?", (mes+"%",)).fetchone()[0]
            egresos  = conn.execute("SELECT COALESCE(SUM(monto),0) FROM finanzas WHERE tipo='Egreso' AND fecha LIKE ?", (mes+"%",)).fetchone()[0]
            todos    = conn.execute("SELECT fecha FROM registros").fetchall()
            
            # KPI de Vencidos (6 meses)
            vencidos = 0
            ahora = datetime.now()
            for (f,) in todos:
                try:
                    fecha_serv = datetime.strptime(f, "%Y-%m-%d")
                    if ahora >= fecha_serv + relativedelta(months=6):
                        vencidos += 1
                except: pass
            
            self.kpis["clientes"].config(text=str(total_clientes))
            self.kpis["hoy"].config(text=str(servicios_hoy))
            self.kpis["ingresos"].config(text=f"${ingresos:,.0f}")
            self.kpis["egresos"].config(text=f"${egresos:,.0f}")
            self.kpis["balance"].config(text=f"${ingresos-egresos:,.0f}")
            self.kpis["vencidos"].config(text=str(vencidos))

            # Actualizar Tabla
            for row in self.tabla.get_children():
                self.tabla.delete(row)
            filas = conn.execute("SELECT nombre,auto,servicio,fecha,costo,estatus FROM registros ORDER BY id DESC LIMIT 15").fetchall()
            conn.close()
            for fila in filas:
                self.tabla.insert("", "end", values=fila)
        except Exception as e:
            print(f"[Error Actualizar Dashboard]: {e}")

    def _recuperar_vencimiento(self):
        from utils.database import get_config
        v = get_config("fecha_vencimiento", "")
        if not v: return None
        try:
            return datetime.fromisoformat(v.replace("Z", "+00:00"))
        except:
            return None

    def _check_timer_vencimiento(self):
        venc = self._recuperar_vencimiento()
        if not venc:
            self.lbl_venc.config(text="Suscripción: Vitalicia", fg=TEXT_MUTED)
            return

        ahora = datetime.now()
        if venc.tzinfo: 
            ahora = datetime.now(venc.tzinfo)
        
        diff = venc - ahora
        segundos = int(diff.total_seconds())

        if segundos <= 0:
            self.lbl_venc.config(text="⚠️ SUSCRIPCIÓN VENCIDA", fg=ERROR)
            return

        dias = segundos // 86400
        horas = (segundos % 86400) // 3600
        mins = (segundos % 3600) // 60

        if dias < 3:
            color = ERROR if dias < 1 else "#f1c40f"
            msg = f"⏱️ Vence en: {horas}h {mins}m" if dias < 1 else f"⏱️ Vence en: {dias}d {horas}h"
            self.lbl_venc.config(text=msg, fg=color)
        else:
            self.lbl_venc.config(text=f"Suscripción activa hasta: {venc.strftime('%d/%m/%Y')}", fg=SUCCESS)

        self.after(60000, self._check_timer_vencimiento)

    def on_show(self):
        self.actualizar()

    def actualizar(self):
        conn = conectar()
        mes = datetime.now().strftime("%Y-%m")
        hoy = datetime.now().strftime("%Y-%m-%d")
        total_clientes = conn.execute("SELECT COUNT(*) FROM registros").fetchone()[0]
        servicios_hoy  = conn.execute("SELECT COUNT(*) FROM registros WHERE fecha=?", (hoy,)).fetchone()[0]
        ingresos = conn.execute("SELECT COALESCE(SUM(monto),0) FROM finanzas WHERE tipo='Ingreso' AND fecha LIKE ?", (mes+"%",)).fetchone()[0]
        egresos  = conn.execute("SELECT COALESCE(SUM(monto),0) FROM finanzas WHERE tipo='Egreso' AND fecha LIKE ?", (mes+"%",)).fetchone()[0]
        todos    = conn.execute("SELECT fecha FROM registros").fetchall()
        conn.close()
        vencidos = 0
        ahora = datetime.now()
        for (f,) in todos:
            try:
                if ahora >= datetime.strptime(f, "%Y-%m-%d") + relativedelta(months=6):
                    vencidos += 1
            except: pass
        self.kpis["clientes"].config(text=str(total_clientes))
        self.kpis["hoy"].config(text=str(servicios_hoy))
        self.kpis["ingresos"].config(text=f"${ingresos:,.0f}")
        self.kpis["egresos"].config(text=f"${egresos:,.0f}")
        self.kpis["balance"].config(text=f"${ingresos-egresos:,.0f}")
        self.kpis["vencidos"].config(text=str(vencidos))
        for row in self.tabla.get_children():
            self.tabla.delete(row)
        conn = conectar()
        filas = conn.execute("SELECT nombre,auto,servicio,fecha,costo,estatus FROM registros ORDER BY id DESC LIMIT 15").fetchall()
        conn.close()
        for fila in filas:
            self.tabla.insert("", "end", values=fila)


class FormularioActivacion(tk.Toplevel):
    def __init__(self, parent):
        super().__init__(parent, bg=BG_DARK)
        self.parent = parent
        self.title("Activación de Producto - TallerPro")
        self.geometry("450x300")
        self.resizable(False, False)
        self.protocol("WM_DELETE_WINDOW", sys.exit)
        
        tk.Label(self, text="BIENVENIDO A TALLERPRO", bg=BG_DARK, fg=ACCENT, 
                 font=("Segoe UI", 16, "bold")).pack(pady=(30, 10))
        tk.Label(self, text="Por favor, ingresa tu clave de activación para continuar:", 
                 bg=BG_DARK, fg=TEXT_MUTED, font=FONT_SMALL).pack()
        
        self.ent_key = tk.Entry(self, bg=BG_CARD, fg=TEXT_PRIMARY, font=("Consolas", 14), 
                                justify="center", bd=1, relief="flat", insertbackground=TEXT_PRIMARY)
        self.ent_key.pack(pady=25, padx=50, fill="x")
        self.ent_key.focus_set()
        
        btn = tk.Button(self, text="ACTIVAR SISTEMA", bg=ACCENT, fg=BG_DARK, relief="flat",
                        command=self.activar, font=("Segoe UI", 10, "bold"), cursor="hand2")
        btn.pack(pady=10, ipadx=20, ipady=5)
        
        self.lbl_status = tk.Label(self, text="", bg=BG_DARK, fg=ERROR, font=FONT_SMALL)
        self.lbl_status.pack()

    def activar(self):
        key = self.ent_key.get().strip().upper()
        if not key: return
        
        self.lbl_status.config(text="Verificando clave...", fg=ACCENT)
        self.update()
        
        try:
            resp = requests.get(f"{API_BASE_URL}/api/v1/license/activate/{key}", timeout=10)
            if resp.status_code == 200:
                data = resp.json()
                set_config("taller_id", str(data["taller_id"]))
                set_config("taller_nombre", data["taller_name"])
                set_config("fecha_vencimiento", data["fecha_vencimiento"])
                messagebox.showinfo("Éxito", f"¡Sistema Activado para {data['taller_name']}!")
                self.destroy()
                self.parent.deiconify()
                python = sys.executable
                os.execl(python, python, *sys.argv)
            else:
                self.lbl_status.config(text="❌ Clave inválida o suspendida.", fg=ERROR)
        except Exception as e:
            self.lbl_status.config(text=f"❌ Error de conexión: {str(e)[:40]}...", fg=ERROR)


class FrameClientes(tk.Frame):
    def __init__(self, parent, app):
        super().__init__(parent, bg=BG_DARK)
        self.app = app
        self._build()

    def _build(self):
        top = tk.Frame(self, bg=BG_DARK, padx=25, pady=14)
        top.pack(fill="x")
        tk.Label(top, text="Gestion de Clientes", bg=BG_DARK,
                 fg=TEXT_PRIMARY, font=FONT_TITLE).pack(side="left")
        btn_frame = tk.Frame(top, bg=BG_DARK)
        btn_frame.pack(side="right")
        boton(btn_frame, "+ Nuevo",        self.nuevo, w=10).pack(side="left", padx=4)
        boton(btn_frame, "Editar",         self.editar, color=BG_INPUT, texto_color=TEXT_PRIMARY, w=8).pack(side="left", padx=4)
        boton(btn_frame, "Borrar",         self.borrar, color=DANGER, texto_color=TEXT_PRIMARY, w=8).pack(side="left", padx=4)
        boton(btn_frame, "Enviar Correo",  self.enviar_correo, color="#2980b9", texto_color=TEXT_PRIMARY, w=12).pack(side="left", padx=4)
        boton(btn_frame, "Notificar",      self.notificar_manual, color="#e67e22", texto_color=TEXT_PRIMARY, w=10).pack(side="left", padx=4)
        boton(btn_frame, "Exportar",       self.exportar, color="#27ae60", texto_color=TEXT_PRIMARY, w=10).pack(side="left", padx=4)
        boton(btn_frame, "Importar",       self.importar, color="#8e44ad", texto_color=TEXT_PRIMARY, w=10).pack(side="left", padx=4)
        search = tk.Frame(self, bg=BG_DARK, padx=25)
        search.pack(fill="x")
        self.buscar_var = tk.StringVar()
        self.buscar_var.trace("w", lambda *a: self.cargar())
        e = tk.Entry(search, textvariable=self.buscar_var,
                     bg=BG_INPUT, fg=TEXT_PRIMARY, font=FONT_BODY,
                     bd=0, insertbackground=ACCENT, width=50)
        e.config(highlightthickness=1, highlightbackground=BORDER, highlightcolor=ACCENT)
        e.pack(side="left", padx=8, pady=8, ipady=5)
        frame_tabla = tk.Frame(self, bg=BG_DARK, padx=25)
        frame_tabla.pack(fill="both", expand=True)
        cols = ("ID","Cliente","Telefono","Correo","Auto","Num Eco","Placas","Servicio","Fecha","Costo","Estatus")
        self.tabla = ttk.Treeview(frame_tabla, columns=cols, show="headings")
        anchos = [40,140,100,160,120,90,90,160,90,80,90]
        for col, w in zip(cols, anchos):
            self.tabla.heading(col, text=col)
            self.tabla.column(col, width=w)
        self.tabla.tag_configure("vencido", background="#3d1515", foreground="#ff6b6b")
        self.tabla.tag_configure("listo",   background="#0d2b1a", foreground="#2ecc71")
        sb = ttk.Scrollbar(frame_tabla, orient="vertical", command=self.tabla.yview)
        self.tabla.configure(yscrollcommand=sb.set)
        self.tabla.pack(side="left", fill="both", expand=True)
        sb.pack(side="right", fill="y")
        self.cargar()

    def on_show(self):
        self.cargar()

    def cargar(self):
        texto = self.buscar_var.get().lower() if hasattr(self, "buscar_var") else ""
        for row in self.tabla.get_children():
            self.tabla.delete(row)
        conn = conectar()
        filas = conn.execute("SELECT * FROM registros ORDER BY id DESC").fetchall()
        conn.close()
        ahora = datetime.now()
        for fila in filas:
            if texto and not any(texto in str(v).lower() for v in fila[1:]):
                continue
            tag = ""
            try:
                venc = datetime.strptime(fila[8], "%Y-%m-%d") + relativedelta(months=6)
                if ahora >= venc:
                    tag = "vencido"
            except: pass
            if fila[10] == "Listo":
                tag = "listo"
            self.tabla.insert("", "end", iid=fila[0], values=fila, tags=(tag,))

    def nuevo(self):
        FormularioCliente(self, self.app)

    def editar(self):
        sel = self.tabla.focus()
        if not sel:
            messagebox.showwarning("Aviso", "Selecciona un registro primero")
            return
        datos = self.tabla.item(sel)["values"]
        FormularioCliente(self, self.app, datos=datos, id_reg=sel)

    def borrar(self):
        sel = self.tabla.focus()
        if not sel: return
        if messagebox.askyesno("Confirmar", "Borrar este registro?"):
            conn = conectar()
            conn.execute("DELETE FROM registros WHERE id=?", (sel,))
            conn.commit()
            conn.close()
            self.cargar()
            self.app.actualizar_dashboard()

    def notificar_manual(self):
        sel = self.tabla.focus()
        if not sel:
            messagebox.showwarning("Aviso", "Selecciona un cliente primero")
            return
        if not AGENTES_OK:
            messagebox.showwarning("Aviso", "Agentes no disponibles")
            return
        ok, msg = self.app.notif_agent.notificar_manual(sel)
        if ok:
            messagebox.showinfo("Notificacion Enviada", msg)
        else:
            messagebox.showwarning("Sin contacto", msg)

    def enviar_correo(self):
        sel = self.tabla.focus()
        if not sel:
            messagebox.showwarning("Aviso", "Selecciona un cliente primero")
            return
        datos = self.tabla.item(sel)["values"]
        correo = datos[3]
        if not correo:
            messagebox.showwarning("Sin correo", "Este cliente no tiene correo registrado")
            return
        EnviarCorreo(self, datos[1], correo, datos[4], datos[7])

    def exportar(self):
        archivo = filedialog.asksaveasfilename(
            defaultextension=".xlsx", filetypes=[("Excel","*.xlsx")])
        if not archivo: return
        conn = conectar()
        filas = conn.execute("SELECT * FROM registros ORDER BY id").fetchall()
        conn.close()
        if not filas:
            messagebox.showwarning("Aviso", "No hay clientes registrados")
            return
        exportar_xlsx_clientes(filas, archivo)
        messagebox.showinfo("Exportado", f"{len(filas)} clientes exportados.\nGuardado en:\n{archivo}")

    def importar(self):
        archivo = filedialog.askopenfilename(
            filetypes=[("Excel","*.xlsx"),("CSV","*.csv")])
        if not archivo: return
        try:
            if archivo.endswith(".xlsx"):
                wb = openpyxl.load_workbook(archivo)
                ws = wb.active
                filas = list(ws.iter_rows(min_row=2, values_only=True))
            else:
                import csv as csvmod
                with open(archivo, encoding="utf-8-sig") as f:
                    filas = list(csvmod.reader(f))[1:]
            conn = conectar()
            importados = 0
            for fila in filas:
                if not any(fila): continue
                try:
                    nombre   = str(fila[1] if len(fila) > 1 else fila[0] or "")
                    telefono = str(fila[2] if len(fila) > 2 else "")
                    correo   = str(fila[3] if len(fila) > 3 else "")
                    auto     = str(fila[4] if len(fila) > 4 else "")
                    num_eco  = str(fila[5] if len(fila) > 5 else "")
                    placas   = str(fila[6] if len(fila) > 6 else "")
                    servicio = str(fila[7] if len(fila) > 7 else "")
                    fecha    = str(fila[8] if len(fila) > 8 else datetime.now().strftime("%Y-%m-%d"))
                    costo    = float(fila[9]) if len(fila) > 9 and fila[9] else 0.0
                    estatus  = str(fila[10] if len(fila) > 10 else "Pendiente")
                    conn.execute("""INSERT INTO registros
                        (nombre,telefono,correo,auto,numero_economico,placas,servicio,fecha,costo,estatus)
                        VALUES (?,?,?,?,?,?,?,?,?,?)""",
                        (nombre,telefono,correo,auto,num_eco,placas,servicio,fecha,costo,estatus))
                    importados += 1
                except: pass
            conn.commit()
            conn.close()
            self.cargar()
            self.app.actualizar_dashboard()
            messagebox.showinfo("Importado", f"{importados} clientes importados correctamente")
        except Exception as ex:
            messagebox.showerror("Error", f"No se pudo importar:\n{ex}")


class FormularioCliente(tk.Toplevel):
    def __init__(self, parent, app, datos=None, id_reg=None):
        super().__init__(parent)
        self.app = app
        self.parent_frame = parent
        self.id_reg = id_reg
        self.title("Nuevo Registro" if not id_reg else "Editar Registro")
        self.geometry("500x780")
        self.configure(bg=BG_CARD)
        self.resizable(True, True)
        self.minsize(480, 600)
        self._build(datos)

    def _build(self, datos):
        canvas = tk.Canvas(self, bg=BG_CARD, bd=0, highlightthickness=0)
        sb = ttk.Scrollbar(self, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=sb.set)
        sb.pack(side="right", fill="y")
        canvas.pack(side="left", fill="both", expand=True)

        frame = tk.Frame(canvas, bg=BG_CARD, padx=30)
        win = canvas.create_window((0,0), window=frame, anchor="nw")

        def _resize(e):
            canvas.configure(scrollregion=canvas.bbox("all"))
            canvas.itemconfig(win, width=canvas.winfo_width())
        frame.bind("<Configure>", _resize)
        canvas.bind("<Configure>", lambda e: canvas.itemconfig(win, width=e.width))
        canvas.bind_all("<MouseWheel>", lambda e: canvas.yview_scroll(int(-1*(e.delta/120)), "units"))

        tk.Label(frame, text="Datos del Servicio", bg=BG_CARD,
                 fg=ACCENT, font=FONT_TITLE).pack(pady=(20,10))
        campos = ["Cliente","Telefono","Correo","Auto","Num. Economico","Placas","Servicio"]
        self.entradas = []
        for campo in campos:
            tk.Label(frame, text=campo, bg=BG_CARD, fg=TEXT_MUTED, font=FONT_SMALL).pack(anchor="w", pady=(8,1))
            e = entrada(frame, width=42)
            e.pack(fill="x", ipady=5)
            self.entradas.append(e)
        row = tk.Frame(frame, bg=BG_CARD)
        row.pack(fill="x", pady=(8,0))
        col1 = tk.Frame(row, bg=BG_CARD)
        col1.pack(side="left", expand=True, fill="x", padx=(0,8))
        col2 = tk.Frame(row, bg=BG_CARD)
        col2.pack(side="left", expand=True, fill="x")
        tk.Label(col1, text="Fecha", bg=BG_CARD, fg=TEXT_MUTED, font=FONT_SMALL).pack(anchor="w")
        self.fecha = DateEntry(col1, date_pattern="yyyy-mm-dd", locale='es')
        self.fecha.pack(fill="x", ipady=4)
        tk.Label(col2, text="Costo ($)", bg=BG_CARD, fg=TEXT_MUTED, font=FONT_SMALL).pack(anchor="w")
        self.costo = entrada(col2, width=15)
        self.costo.pack(fill="x", ipady=5)
        tk.Label(frame, text="Estatus", bg=BG_CARD, fg=TEXT_MUTED, font=FONT_SMALL).pack(anchor="w", pady=(8,1))
        self.estatus = ttk.Combobox(frame, values=["Pendiente","En proceso","Listo","Entregado"], state="readonly", font=FONT_BODY)
        self.estatus.set("Pendiente")
        self.estatus.pack(fill="x", ipady=4)
        if datos:
            vals = list(datos)
            for i, e in enumerate(self.entradas):
                e.insert(0, str(vals[i+1]))
            try: self.fecha.set_date(vals[8])
            except: pass
            self.costo.insert(0, str(vals[9] or ""))
            self.estatus.set(vals[10] or "Pendiente")
        boton(frame, "Guardar", self.guardar, w=20).pack(pady=20)

    def guardar(self):
        vals = [e.get() for e in self.entradas]
        fecha_val = self.fecha.get()
        try: costo_val = float(self.costo.get() or 0)
        except: costo_val = 0.0
        estatus_val = self.estatus.get()
        conn = conectar()
        if self.id_reg:
            conn.execute("""UPDATE registros SET nombre=?,telefono=?,correo=?,auto=?,
                numero_economico=?,placas=?,servicio=?,fecha=?,costo=?,estatus=?
                WHERE id=?""", (*vals, fecha_val, costo_val, estatus_val, self.id_reg))
        else:
            conn.execute("""INSERT INTO registros
                (nombre,telefono,correo,auto,numero_economico,placas,servicio,fecha,costo,estatus)
                VALUES (?,?,?,?,?,?,?,?,?,?)""", (*vals, fecha_val, costo_val, estatus_val))
        conn.commit()
        conn.close()
        self.parent_frame.cargar()
        self.app.actualizar_dashboard()
        self.destroy()


class FrameFinanzas(tk.Frame):
    def __init__(self, parent, app):
        super().__init__(parent, bg=BG_DARK)
        self.app = app
        self._build()

    def _build(self):
        top = tk.Frame(self, bg=BG_DARK, padx=25, pady=14)
        top.pack(fill="x")
        tk.Label(top, text="Control de Finanzas", bg=BG_DARK,
                 fg=TEXT_PRIMARY, font=FONT_TITLE).pack(side="left")
        boton(top, "+ Agregar", self.nuevo, w=12).pack(side="right", padx=4)
        boton(top, "Borrar", self.borrar, color=DANGER, texto_color=TEXT_PRIMARY, w=10).pack(side="right", padx=4)
        resumen = tk.Frame(self, bg=BG_DARK, padx=25)
        resumen.pack(fill="x")
        self.lbl_ing = tk.Label(resumen, text="Ingresos: $0", bg=BG_CARD,
                                fg=SUCCESS, font=("Segoe UI",13,"bold"), padx=20, pady=10)
        self.lbl_ing.pack(side="left", padx=6)
        self.lbl_eg = tk.Label(resumen, text="Egresos: $0", bg=BG_CARD,
                               fg=DANGER, font=("Segoe UI",13,"bold"), padx=20, pady=10)
        self.lbl_eg.pack(side="left", padx=6)
        self.lbl_bal = tk.Label(resumen, text="Balance: $0", bg=BG_CARD,
                                fg=ACCENT, font=("Segoe UI",13,"bold"), padx=20, pady=10)
        self.lbl_bal.pack(side="left", padx=6)
        frame_t = tk.Frame(self, bg=BG_DARK, padx=25, pady=10)
        frame_t.pack(fill="both", expand=True)
        cols = ("ID","Tipo","Concepto","Monto","Fecha","Notas")
        self.tabla = ttk.Treeview(frame_t, columns=cols, show="headings")
        anchos = [40,80,200,100,100,300]
        for col, w in zip(cols, anchos):
            self.tabla.heading(col, text=col)
            self.tabla.column(col, width=w)
        self.tabla.tag_configure("ingreso", foreground=SUCCESS)
        self.tabla.tag_configure("egreso",  foreground=DANGER)
        sb = ttk.Scrollbar(frame_t, orient="vertical", command=self.tabla.yview)
        self.tabla.configure(yscrollcommand=sb.set)
        self.tabla.pack(side="left", fill="both", expand=True)
        sb.pack(side="right", fill="y")
        self.cargar()

    def on_show(self):
        self.cargar()

    def cargar(self):
        for row in self.tabla.get_children():
            self.tabla.delete(row)
        conn = conectar()
        filas = conn.execute("SELECT * FROM finanzas ORDER BY id DESC").fetchall()
        mes = datetime.now().strftime("%Y-%m")
        ing = conn.execute("SELECT COALESCE(SUM(monto),0) FROM finanzas WHERE tipo='Ingreso' AND fecha LIKE ?", (mes+"%",)).fetchone()[0]
        eg  = conn.execute("SELECT COALESCE(SUM(monto),0) FROM finanzas WHERE tipo='Egreso' AND fecha LIKE ?", (mes+"%",)).fetchone()[0]
        conn.close()
        for fila in filas:
            tag = "ingreso" if fila[1] == "Ingreso" else "egreso"
            self.tabla.insert("", "end", iid=fila[0], values=fila, tags=(tag,))
        self.lbl_ing.config(text=f"Ingresos del Mes: ${ing:,.2f}")
        self.lbl_eg.config(text=f"Egresos del Mes: ${eg:,.2f}")
        bal = ing - eg
        self.lbl_bal.config(text=f"Balance: ${bal:,.2f}", fg=SUCCESS if bal >= 0 else DANGER)
        self.app.actualizar_dashboard()

    def nuevo(self):
        FormularioFinanza(self)

    def borrar(self):
        sel = self.tabla.focus()
        if not sel: return
        if messagebox.askyesno("Confirmar", "Borrar este movimiento?"):
            conn = conectar()
            conn.execute("DELETE FROM finanzas WHERE id=?", (sel,))
            conn.commit()
            conn.close()
            self.cargar()


class FormularioFinanza(tk.Toplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.parent_frame = parent
        self.title("Nuevo Movimiento")
        self.geometry("420x380")
        self.configure(bg=BG_CARD)
        self.resizable(False, False)
        self._build()

    def _build(self):
        tk.Label(self, text="Nuevo Movimiento", bg=BG_CARD,
                 fg=ACCENT, font=FONT_TITLE).pack(pady=(20,10))
        frame = tk.Frame(self, bg=BG_CARD, padx=30)
        frame.pack(fill="x")
        tk.Label(frame, text="Tipo", bg=BG_CARD, fg=TEXT_MUTED, font=FONT_SMALL).pack(anchor="w")
        self.tipo = ttk.Combobox(frame, values=["Ingreso","Egreso"], state="readonly", font=FONT_BODY)
        self.tipo.set("Ingreso")
        self.tipo.pack(fill="x", ipady=4)
        tk.Label(frame, text="Concepto", bg=BG_CARD, fg=TEXT_MUTED, font=FONT_SMALL).pack(anchor="w", pady=(8,1))
        self.concepto = entrada(frame, width=42)
        self.concepto.pack(fill="x", ipady=5)
        tk.Label(frame, text="Monto ($)", bg=BG_CARD, fg=TEXT_MUTED, font=FONT_SMALL).pack(anchor="w", pady=(8,1))
        self.monto = entrada(frame, width=42)
        self.monto.pack(fill="x", ipady=5)
        tk.Label(frame, text="Fecha", bg=BG_CARD, fg=TEXT_MUTED, font=FONT_SMALL).pack(anchor="w", pady=(8,1))
        self.fecha = DateEntry(frame, date_pattern="yyyy-mm-dd", locale='es')
        self.fecha.pack(fill="x", ipady=4)
        tk.Label(frame, text="Notas", bg=BG_CARD, fg=TEXT_MUTED, font=FONT_SMALL).pack(anchor="w", pady=(8,1))
        self.notas = entrada(frame, width=42)
        self.notas.pack(fill="x", ipady=5)
        boton(self, "Guardar", self.guardar, w=20).pack(pady=20)

    def guardar(self):
        try: monto = float(self.monto.get() or 0)
        except: monto = 0.0
        conn = conectar()
        conn.execute("INSERT INTO finanzas (tipo,concepto,monto,fecha,notas) VALUES (?,?,?,?,?)",
                     (self.tipo.get(), self.concepto.get(), monto, self.fecha.get(), self.notas.get()))
        conn.commit()
        conn.close()
        self.parent_frame.cargar()
        self.destroy()


class FrameReportes(tk.Frame):
    def __init__(self, parent, app):
        super().__init__(parent, bg=BG_DARK)
        self.app = app
        self._build()

    def _build(self):
        top = tk.Frame(self, bg=BG_DARK, padx=25, pady=14)
        top.pack(fill="x")
        tk.Label(top, text="Reportes", bg=BG_DARK, fg=TEXT_PRIMARY, font=FONT_TITLE).pack(side="left")
        content = tk.Frame(self, bg=BG_DARK, padx=25)
        content.pack(fill="both", expand=True)
        cards = [
            ("Reporte Mensual de Finanzas", "Excel con ingresos, egresos y balance del mes", self.reporte_finanzas),
            ("Reporte de Clientes", "Lista completa de clientes con servicios y costos", self.reporte_clientes),
            ("Reporte de Servicios Vencidos", "Clientes con servicio vencido hace mas de 6 meses", self.reporte_vencidos),
        ]
        for titulo, desc, cmd in cards:
            card = tk.Frame(content, bg=BG_CARD, padx=20, pady=15,
                            highlightthickness=1, highlightbackground=BORDER)
            card.pack(fill="x", pady=8)
            tk.Label(card, text=titulo, bg=BG_CARD, fg=ACCENT, font=FONT_HEADER).pack(anchor="w")
            tk.Label(card, text=desc, bg=BG_CARD, fg=TEXT_MUTED, font=FONT_SMALL).pack(anchor="w", pady=4)
            boton(card, "Generar", cmd, w=14).pack(anchor="w", pady=6)

    def reporte_finanzas(self):
        archivo = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel","*.xlsx")])
        if not archivo: return
        conn = conectar()
        mes = datetime.now().strftime("%Y-%m")
        filas = conn.execute("SELECT * FROM finanzas WHERE fecha LIKE ? ORDER BY fecha", (mes+"%",)).fetchall()
        conn.close()
        generar_reporte_finanzas(filas, archivo, mes)
        messagebox.showinfo("Reporte generado", f"Guardado en:\n{archivo}")

    def reporte_clientes(self):
        archivo = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel","*.xlsx")])
        if not archivo: return
        conn = conectar()
        filas = conn.execute("SELECT * FROM registros ORDER BY nombre").fetchall()
        conn.close()
        exportar_xlsx_clientes(filas, archivo)
        messagebox.showinfo("Reporte generado", f"Guardado en:\n{archivo}")

    def reporte_vencidos(self):
        archivo = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel","*.xlsx")])
        if not archivo: return
        conn = conectar()
        filas = conn.execute("SELECT * FROM registros").fetchall()
        conn.close()
        ahora = datetime.now()
        vencidos = [f for f in filas if _esta_vencido(f[8], ahora)]
        exportar_xlsx_clientes(vencidos, archivo, titulo="Servicios Vencidos")
        messagebox.showinfo("Reporte generado", f"{len(vencidos)} servicios vencidos.\nGuardado en:\n{archivo}")


def _esta_vencido(fecha_str, ahora):
    try:
        return ahora >= datetime.strptime(fecha_str, "%Y-%m-%d") + relativedelta(months=6)
    except:
        return False


class FrameConfig(tk.Frame):
    def __init__(self, parent, app):
        super().__init__(parent, bg=BG_DARK)
        self.app = app
        self._build()

    def _build(self):
        top = tk.Frame(self, bg=BG_DARK, padx=25, pady=14)
        top.pack(fill="x")
        tk.Label(top, text="Configuracion", bg=BG_DARK,
                 fg=TEXT_PRIMARY, font=FONT_TITLE).pack(side="left")

        canvas = tk.Canvas(self, bg=BG_DARK, bd=0, highlightthickness=0)
        sb = ttk.Scrollbar(self, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=sb.set)
        sb.pack(side="right", fill="y")
        canvas.pack(side="left", fill="both", expand=True, padx=25)

        frame = tk.Frame(canvas, bg=BG_CARD, padx=30, pady=20)
        win = canvas.create_window((0,0), window=frame, anchor="nw")

        def _resize(e):
            canvas.configure(scrollregion=canvas.bbox("all"))
            canvas.itemconfig(win, width=canvas.winfo_width())
        frame.bind("<Configure>", _resize)
        canvas.bind("<Configure>", lambda e: canvas.itemconfig(win, width=e.width))
        canvas.bind_all("<MouseWheel>", lambda e: canvas.yview_scroll(int(-1*(e.delta/120)), "units"))

        campos = [
            ("Correo SMTP (Gmail)",      "smtp_email"),
            ("Contrasena de App Gmail",  "smtp_pass"),
            ("Nombre del Taller",        "taller_nombre"),
            ("Telefono del Taller",      "taller_tel"),
            ("Ubicacion",                "ubicacion"),
            ("Horario de Atencion",      "horario"),
            ("Servicios y precios",      "servicios_info"),
            ("WhatsApp Instance ID",     "whatsapp_instance"),
            ("WhatsApp Token",           "whatsapp_token"),
            ("Claude API Key (opcional)","claude_api_key"),
        ]
        self.entradas = {}
        for label_txt, key in campos:
            tk.Label(frame, text=label_txt, bg=BG_CARD, fg=TEXT_MUTED,
                     font=FONT_SMALL).pack(anchor="w", pady=(8,1))
            e = entrada(frame, width=55)
            e.pack(fill="x", ipady=5)
            self.entradas[key] = e

        boton(frame, "Guardar Config", self.guardar, w=20).pack(pady=20)
        self.cargar_config()

    def cargar_config(self):
        conn = conectar()
        for key, e in self.entradas.items():
            row = conn.execute("SELECT valor FROM config WHERE clave=?", (key,)).fetchone()
            if row:
                e.delete(0, "end")
                e.insert(0, row[0])
        conn.close()

    def guardar(self):
        conn = conectar()
        for key, e in self.entradas.items():
            conn.execute("INSERT OR REPLACE INTO config (clave,valor) VALUES (?,?)", (key, e.get()))
        conn.commit()
        conn.close()
        messagebox.showinfo("OK", "Configuracion guardada")


class EnviarCorreo(tk.Toplevel):
    def __init__(self, parent, nombre, correo, auto, servicio):
        super().__init__(parent)
        self.title("Enviar Correo")
        self.geometry("500x400")
        self.configure(bg=BG_CARD)
        self.nombre = nombre
        self.correo = correo
        self.auto = auto
        self.servicio = servicio
        self._build()

    def _build(self):
        tk.Label(self, text="Enviar Correo al Cliente", bg=BG_CARD,
                 fg=ACCENT, font=FONT_TITLE).pack(pady=(20,10))
        frame = tk.Frame(self, bg=BG_CARD, padx=30)
        frame.pack(fill="x")
        tk.Label(frame, text=f"Para: {self.correo}", bg=BG_CARD, fg=TEXT_MUTED, font=FONT_SMALL).pack(anchor="w")
        tk.Label(frame, text="Asunto", bg=BG_CARD, fg=TEXT_MUTED, font=FONT_SMALL).pack(anchor="w", pady=(8,1))
        self.asunto = entrada(frame, width=50)
        self.asunto.insert(0, f"Tu vehiculo {self.auto} esta listo")
        self.asunto.pack(fill="x", ipady=5)
        tk.Label(frame, text="Mensaje", bg=BG_CARD, fg=TEXT_MUTED, font=FONT_SMALL).pack(anchor="w", pady=(8,1))
        self.mensaje = tk.Text(frame, bg=BG_INPUT, fg=TEXT_PRIMARY, font=FONT_BODY, bd=0, height=8, insertbackground=ACCENT)
        self.mensaje.pack(fill="x")
        self.mensaje.insert("1.0", f"Estimado/a {self.nombre},\n\nSu vehiculo {self.auto} ya esta listo.\n\nServicio: {self.servicio}\n\nGracias por su preferencia.")
        boton(self, "Enviar", self.enviar, w=20).pack(pady=15)

    def enviar(self):
        conn = conectar()
        smtp_email = conn.execute("SELECT valor FROM config WHERE clave='smtp_email'").fetchone()
        smtp_pass  = conn.execute("SELECT valor FROM config WHERE clave='smtp_pass'").fetchone()
        conn.close()
        if not smtp_email or not smtp_pass:
            messagebox.showwarning("Config", "Configura el correo SMTP primero")
            return
        try:
            msg = MIMEMultipart()
            msg["From"]    = smtp_email[0]
            msg["To"]      = self.correo
            msg["Subject"] = self.asunto.get()
            msg.attach(MIMEText(self.mensaje.get("1.0","end"), "plain"))
            server = smtplib.SMTP("smtp.gmail.com", 587)
            server.starttls()
            server.login(smtp_email[0], smtp_pass[0])
            server.send_message(msg)
            server.quit()
            messagebox.showinfo("OK", f"Correo enviado a {self.correo}")
            self.destroy()
        except Exception as ex:
            messagebox.showerror("Error", f"No se pudo enviar:\n{ex}")


def exportar_xlsx_clientes(filas, archivo, titulo="Clientes"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = titulo
    amarillo = PatternFill("solid", fgColor="E8A020")
    bold_b   = Font(bold=True, color="000000")
    center   = Alignment(horizontal="center", vertical="center")
    headers  = ["ID","Cliente","Telefono","Correo","Auto","Num Eco","Placas","Servicio","Fecha","Costo","Estatus"]
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = amarillo
        cell.font = bold_b
        cell.alignment = center
    for fila in filas:
        ws.append(list(fila))
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)
    wb.save(archivo)


def generar_reporte_finanzas(filas, archivo, mes):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Finanzas {mes}"
    amarillo = PatternFill("solid", fgColor="E8A020")
    bold_b   = Font(bold=True, color="000000")
    center   = Alignment(horizontal="center")
    headers  = ["ID","Tipo","Concepto","Monto","Fecha","Notas"]
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = amarillo
        cell.font = bold_b
        cell.alignment = center
    ing = eg = 0
    for fila in filas:
        ws.append(list(fila))
        if fila[1] == "Ingreso": ing += fila[3]
        else: eg += fila[3]
    ws.append([])
    ws.append(["","","INGRESOS", ing])
    ws.append(["","","EGRESOS",  eg])
    ws.append(["","","BALANCE",  ing - eg])
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)
    wb.save(archivo)



class FrameLogs(tk.Frame):
    def __init__(self, parent, app):
        super().__init__(parent, bg=BG_DARK)
        self.app = app
        self._build()

    def _build(self):
        top = tk.Frame(self, bg=BG_DARK, padx=25, pady=14)
        top.pack(fill="x")
        tk.Label(top, text="Logs del Sistema", bg=BG_DARK,
                 fg=TEXT_PRIMARY, font=FONT_TITLE).pack(side="left")
        boton(top, "Actualizar", self.cargar, color=BG_INPUT,
              texto_color=TEXT_PRIMARY, w=12).pack(side="right", padx=4)
        boton(top, "Limpiar Logs", self.limpiar, color=DANGER,
              texto_color=TEXT_PRIMARY, w=12).pack(side="right", padx=4)

        frame_t = tk.Frame(self, bg=BG_DARK, padx=25)
        frame_t.pack(fill="both", expand=True)

        cols = ("Fecha","Nivel","Modulo","Mensaje")
        self.tabla = ttk.Treeview(frame_t, columns=cols, show="headings")
        anchos = [140, 60, 150, 500]
        for col, w in zip(cols, anchos):
            self.tabla.heading(col, text=col)
            self.tabla.column(col, width=w)
        self.tabla.tag_configure("ERROR", foreground=DANGER)
        self.tabla.tag_configure("INFO",  foreground=SUCCESS)
        self.tabla.tag_configure("WARN",  foreground=ACCENT)
        sb = ttk.Scrollbar(frame_t, orient="vertical", command=self.tabla.yview)
        self.tabla.configure(yscrollcommand=sb.set)
        self.tabla.pack(side="left", fill="both", expand=True)
        sb.pack(side="right", fill="y")
        self.cargar()

    def on_show(self):
        self.cargar()

    def cargar(self):
        for row in self.tabla.get_children():
            self.tabla.delete(row)
        try:
            conn = conectar()
            filas = conn.execute(
                "SELECT fecha, nivel, modulo, mensaje FROM logs ORDER BY id DESC LIMIT 200"
            ).fetchall()
            conn.close()
            for fila in filas:
                tag = fila[1] if fila[1] in ("ERROR","INFO","WARN") else ""
                self.tabla.insert("", "end", values=fila, tags=(tag,))
        except: pass

    def limpiar(self):
        if messagebox.askyesno("Confirmar", "Borrar todos los logs?"):
            conn = conectar()
            conn.execute("DELETE FROM logs")
            conn.commit()
            conn.close()
            self.cargar()

class FrameAgentes(tk.Frame):
    def __init__(self, parent, app):
        super().__init__(parent, bg=BG_DARK)
        self.app = app
        self._build()

    def _build(self):
        top = tk.Frame(self, bg=BG_DARK, padx=25, pady=14)
        top.pack(fill="x")
        tk.Label(top, text="Agentes IA — Command Center", bg=BG_DARK, fg=TEXT_PRIMARY, font=FONT_TITLE).pack(side="left")

        content = tk.Frame(self, bg=BG_DARK, padx=25)
        content.pack(fill="both", expand=True)

        console_frame = tk.Frame(content, bg=BG_DARK, pady=10)
        console_frame.pack(fill="both", expand=True)

        tk.Label(console_frame, text="Consola de Pensamiento de Agentes (Tiempo Real)", 
                 bg=BG_DARK, fg=ACCENT, font=FONT_HEADER).pack(anchor="w", pady=(0,5))
        
        self.console = tk.Text(console_frame, bg="#0d1117", fg="#c9d1d9",
                               font=("Consolas", 10), bd=0, height=12,
                               padx=15, pady=15, state="disabled")
        sb_c = ttk.Scrollbar(console_frame, orient="vertical", command=self.console.yview)
        self.console.configure(yscrollcommand=sb_c.set)
        self.console.pack(side="left", fill="both", expand=True)
        sb_c.pack(side="right", fill="y")
        
        board_label = tk.Label(content, text="Oportunidades Detectadas por Lead Detective", 
                               bg=BG_DARK, fg=SUCCESS, font=FONT_HEADER)
        board_label.pack(anchor="w", pady=(20,5))
        
        cols_p = ("ID", "Nombre", "Vehiculo", "Interes", "Fecha")
        self.tabla_p = ttk.Treeview(content, columns=cols_p, show="headings", height=5)
        for col in cols_p:
            self.tabla_p.heading(col, text=col)
            self.tabla_p.column(col, width=100)
        self.tabla_p.pack(fill="x", pady=(0, 20))
        
        boton(content, "Importar Excel Contador (4 tabs)", self.importar_excel_contador,
              color=ACCENT, w=22).pack(anchor="w")

        self.cargar_logs()
        self.cargar_prospectos()
        self._set_auto_refresh()

    def _set_auto_refresh(self):
        self.cargar_logs()
        self.cargar_prospectos()
        self.after(5000, self._set_auto_refresh)

    def cargar_logs(self):
        try:
            conn = conectar()
            rows = conn.execute("SELECT agente, mensaje, fecha FROM agentes_logs ORDER BY id DESC LIMIT 50").fetchall()
            conn.close()
            
            self.console.config(state="normal")
            self.console.delete("1.0", "end")
            for r in reversed(rows):
                try: t = r[2][11:19]
                except: t = "--:--:--"
                prefix = f"[{t}] {r[0]}: "
                self.console.insert("end", prefix, "tag_prefix")
                self.console.insert("end", f"{r[1]}\n", "tag_msg")
            
            self.console.tag_config("tag_prefix", foreground=ACCENT, font=("Consolas", 10, "bold"))
            self.console.tag_config("tag_msg",    foreground="#c9d1d9")
            self.console.config(state="disabled")
            self.console.see("end")
        except: pass

    def cargar_prospectos(self):
        try:
            for r in self.tabla_p.get_children(): self.tabla_p.delete(r)
            conn = conectar()
            rows = conn.execute("SELECT id, nombre, auto, interes, fecha FROM prospectos ORDER BY id DESC").fetchall()
            conn.close()
            for r in rows:
                self.tabla_p.insert("", "end", values=r)
        except: pass

    def importar_excel_contador(self):
        archivo = filedialog.askopenfilename(
            title="Selecciona el Excel del Contador",
            filetypes=[("Excel","*.xlsx")])
        if not archivo: return
        try:
            from agentes.contador_import_agent import importar_excel_contador
            resultados = importar_excel_contador(archivo)
            if not resultados:
                messagebox.showerror("Error", "No se pudo importar el archivo")
                return
            resumen = "IMPORTACION COMPLETADA\n" + "="*35 + "\n"
            total_imp = 0
            total_dup = 0
            nombres = {
                "relacion_carros": "Relacion de Carros",
                "informe_diario":  "Informe Diario",
                "mano_obra":       "Mano de Obra",
                "nomina":          "Nomina",
            }
            for key, nombre in nombres.items():
                if key in resultados:
                    imp = resultados[key].get("importados", 0)
                    dup = resultados[key].get("duplicados", 0)
                    total_imp += imp
                    total_dup += dup
                    resumen += f"\n{nombre}:\n  Nuevos: {imp}  |  Duplicados: {dup}\n"
            resumen += f"\n{'='*35}\nTOTAL IMPORTADOS: {total_imp}\nDUPLICADOS OMITIDOS: {total_dup}"
            messagebox.showinfo("Importacion Completada", resumen)
            if "FrameFinanzas" in self.app.frames:
                self.app.frames["FrameFinanzas"].cargar()
            if "FrameClientes" in self.app.frames:
                self.app.frames["FrameClientes"].cargar()
            self.app.actualizar_dashboard()
        except Exception as ex:
            messagebox.showerror("Error", f"Error importando:\n{ex}")

    def activar_vigilancia(self):
        carpeta = filedialog.askdirectory(title="Selecciona carpeta a vigilar")
        if not carpeta: return
        self.app.finance_agent.iniciar_vigilancia(carpeta)
        messagebox.showinfo("Vigilancia Activa",
                            f"Monitoreando:\n{carpeta}\n\nCualquier Excel o CSV nuevo sera importado automaticamente.")


class FrameChatbot(tk.Frame):
    def __init__(self, parent, app):
        super().__init__(parent, bg=BG_DARK)
        self.app = app
        self.historial = []
        self.esperando = None
        self.datos_cita = {}
        self._build()

    def _build(self):
        top = tk.Frame(self, bg=BG_DARK, padx=25, pady=14)
        top.pack(fill="x")
        tk.Label(top, text="Chatbot de Ventas", bg=BG_DARK,
                 fg=TEXT_PRIMARY, font=FONT_TITLE).pack(side="left")
        boton(top, "Limpiar Chat", self.limpiar, color=BG_INPUT,
              texto_color=TEXT_PRIMARY, w=12).pack(side="right")
        tk.Label(top, text="Simulador — prueba como responde el agente a tus clientes",
                 bg=BG_DARK, fg=TEXT_MUTED, font=FONT_SMALL).pack(side="right", padx=10)

        chat_frame = tk.Frame(self, bg=BG_DARK, padx=25)
        chat_frame.pack(fill="both", expand=True)

        self.chat_text = tk.Text(chat_frame, bg=BG_CARD, fg=TEXT_PRIMARY,
                                  font=FONT_BODY, bd=0, state="disabled",
                                  wrap="word", height=20)
        sb = ttk.Scrollbar(chat_frame, orient="vertical", command=self.chat_text.yview)
        self.chat_text.configure(yscrollcommand=sb.set)
        self.chat_text.pack(side="left", fill="both", expand=True)
        sb.pack(side="right", fill="y")

        input_frame = tk.Frame(self, bg=BG_DARK, padx=25, pady=10)
        input_frame.pack(fill="x")
        self.entrada_msg = entrada(input_frame, width=70)
        self.entrada_msg.pack(side="left", fill="x", expand=True, ipady=8, padx=(0,10))
        self.entrada_msg.bind("<Return>", lambda e: self.enviar())
        boton(input_frame, "Enviar", self.enviar, w=10).pack(side="left")

        self._agregar_msg("Agente", "Hola! Bienvenido. Puedo ayudarte con precios, citas, horarios y servicios. En que te puedo ayudar?", ACCENT)

    def _agregar_msg(self, quien, texto, color=TEXT_PRIMARY):
        self.chat_text.config(state="normal")
        self.chat_text.insert("end", f"\n{quien}:\n", "label")
        self.chat_text.insert("end", f"{texto}\n", "msg")
        self.chat_text.tag_config("label", foreground=color, font=FONT_HEADER)
        self.chat_text.tag_config("msg",   foreground=TEXT_PRIMARY, font=FONT_BODY)
        self.chat_text.config(state="disabled")
        self.chat_text.see("end")

    def enviar(self):
        msg = self.entrada_msg.get().strip()
        if not msg: return
        self.entrada_msg.delete(0, "end")
        self._agregar_msg("Cliente", msg, "#2980b9")

        if self.esperando == "fecha_cita":
            self.datos_cita["fecha"] = msg
            self.esperando = "nombre_cita"
            self._agregar_msg("Agente",
                f"Perfecto, el {msg}. Para confirmar necesito tu nombre completo.", SUCCESS)
            return

        if self.esperando == "nombre_cita":
            self.datos_cita["nombre"] = msg
            self.esperando = "tel_cita"
            self._agregar_msg("Agente",
                f"Gracias {msg}. Y tu numero de telefono para confirmarte?", SUCCESS)
            return

        if self.esperando == "tel_cita":
            self.datos_cita["tel"] = msg
            self.esperando = None
            serv = self.app.chat_agent.contexto_servicio or "servicio"
            taller = get_config("taller_nombre", "Taller Pro")
            self._agregar_msg("Agente",
                f"Listo! Cita agendada para {self.datos_cita.get('nombre')} "
                f"el {self.datos_cita.get('fecha')} para revision de {serv}. "
                f"Te esperamos en {taller}. "
                f"Te confirmamos al {msg}. Hasta pronto!", SUCCESS)
            self.datos_cita = {}
            return

        msg_lower = msg.lower()
        afirmaciones = ["si","sí","ok","claro","dale","andale","va","perfecto","por favor","porfavor","quiero","me gustaria","adelante","yes"]
        tiene_contexto = AGENTES_OK and self.app.chat_agent.contexto_servicio
        
        if tiene_contexto and any(p == msg_lower.strip() or msg_lower.strip().startswith(p) for p in afirmaciones):
            self.esperando = "fecha_cita"
            serv = self.app.chat_agent.contexto_servicio
            self._agregar_msg("Agente",
                f"Con gusto te agendo la cita para {serv}. Que dia y hora te viene mejor?", SUCCESS)
            return

        if any(p in msg_lower for p in ["cita","agendar","reservar","cuando puedo",
                                         "mañana","hoy","lunes","martes","miercoles",
                                         "jueves","viernes","sabado"]):
            self.esperando = "fecha_cita"
            serv = self.app.chat_agent.contexto_servicio if AGENTES_OK else None
            if serv:
                self._agregar_msg("Agente",
                    f"Con gusto te agendo la cita para el {serv}. Que dia y hora te viene mejor?", SUCCESS)
            else:
                self._agregar_msg("Agente",
                    "Con gusto te agendo una cita. Que dia y hora te viene mejor?", SUCCESS)
            return

        respuesta = self.app.chat_agent.responder(msg)
        self._agregar_msg("Agente", respuesta, SUCCESS)

    def limpiar(self):
        self.chat_text.config(state="normal")
        self.chat_text.delete("1.0", "end")
        self.chat_text.config(state="disabled")
        self.historial = []
        self.esperando = None
        self.datos_cita = {}
        if AGENTES_OK:
            self.app.chat_agent.resetear_contexto()
        self._agregar_msg("Agente", "Chat reiniciado. En que te puedo ayudar?", ACCENT)


if __name__ == "__main__":
    import hashlib

    def hash_pass(p):
        return hashlib.sha256(p.encode()).hexdigest()

    def verificar_login(usuario, password):
        conn = conectar()
        row = conn.execute("SELECT valor FROM config WHERE clave='app_password'").fetchone()
        usr = conn.execute("SELECT valor FROM config WHERE clave='app_usuario'").fetchone()
        conn.close()
        u_db = usr[0] if usr else "admin"
        p_db = row[0] if row else hash_pass("1234")
        return usuario == u_db and hash_pass(password) == p_db

    class LoginWindow(tk.Tk):
        def __init__(self):
            super().__init__()
            self.title("TallerPro - Acceso")
            self.geometry("420x500")
            self.configure(bg=BG_DARK)
            self.resizable(False, False)
            self.resultado = False
            self._build()
            try:
                ico = os.path.join(os.path.dirname(os.path.abspath(__file__)), "icono.ico")
                if os.path.exists(ico): self.iconbitmap(ico)
            except: pass
            self.entry_pass.focus()

        def _build(self):
            tk.Label(self, text="TALLER PRO", bg=BG_DARK, fg=ACCENT,
                     font=("Segoe UI", 24, "bold")).pack(pady=(60, 5))
            tk.Label(self, text="Sistema de Gestion Inteligente", bg=BG_DARK,
                     fg=TEXT_MUTED, font=FONT_SMALL).pack(pady=(0, 30))
            frame = tk.Frame(self, bg=BG_DARK, padx=50)
            frame.pack(fill="x")
            tk.Label(frame, text="Usuario", bg=BG_DARK, fg=TEXT_MUTED, font=FONT_SMALL).pack(anchor="w")
            self.entry_user = tk.Entry(frame, bg=BG_INPUT, fg=TEXT_PRIMARY, font=FONT_BODY,
                                       bd=0, insertbackground=ACCENT)
            self.entry_user.config(highlightthickness=1, highlightbackground=BORDER, highlightcolor=ACCENT)
            self.entry_user.pack(fill="x", ipady=8, pady=(2,12))
            self.entry_user.insert(0, "admin")
            tk.Label(frame, text="Contrasena", bg=BG_DARK, fg=TEXT_MUTED, font=FONT_SMALL).pack(anchor="w")
            self.entry_pass = tk.Entry(frame, bg=BG_INPUT, fg=TEXT_PRIMARY, font=FONT_BODY,
                                       bd=0, insertbackground=ACCENT, show="*")
            self.entry_pass.config(highlightthickness=1, highlightbackground=BORDER, highlightcolor=ACCENT)
            self.entry_pass.pack(fill="x", ipady=8, pady=(2,20))
            self.entry_pass.bind("<Return>", lambda e: self.login())
            self.lbl_error = tk.Label(frame, text="", bg=BG_DARK, fg=DANGER, font=FONT_SMALL)
            self.lbl_error.pack()
            tk.Button(frame, text="Entrar", command=self.login,
                      bg=ACCENT, fg=BG_DARK, font=FONT_HEADER, bd=0, relief="flat",
                      padx=20, pady=10, cursor="hand2",
                      activebackground=ACCENT_DARK).pack(fill="x", pady=10)
            tk.Label(self, text="Contrasena por defecto: 1234", bg=BG_DARK,
                     fg=TEXT_MUTED, font=("Segoe UI", 8)).pack(side="bottom", pady=10)

        def login(self):
            u = self.entry_user.get().strip()
            p = self.entry_pass.get().strip()
            if verificar_login(u, p):
                self.resultado = True
                self.destroy()
            else:
                self.lbl_error.config(text="Usuario o contrasena incorrectos")
                self.entry_pass.delete(0, "end")

    crear_tablas()
    if AGENTES_OK:
        _crear_tablas_ext()

    # --- REGLA SaaS #1: ACTIVACIÓN PRIMERO ---
    tid = get_config("taller_id", "")
    if not tid:
        root = tk.Tk()
        root.withdraw()
        FormularioActivacion(root)
        root.mainloop()
        sys.exit() # El formulario reiniciará la app tras activar

    # --- REGLA SaaS #2: LOGIN DESPUÉS ---
    login = LoginWindow()
    login.mainloop()

    if login.resultado:
        app = TallerPro()
        app.mainloop()