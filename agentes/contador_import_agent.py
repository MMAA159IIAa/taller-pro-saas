import openpyxl
import hashlib
import sqlite3
import os
import sys
from datetime import datetime

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from utils.database import conectar
from utils.logger import info, error, warn

EMPLEADOS_CONOCIDOS = {
    "CARLOS DANIEL GUILLEN MONTAÑO", "FIDEL ENRIQUE LARA LOPEZ",
    "GUADALUPE LOREDO BELTRAN", "JESUS ENRIQUE RAMIREZ ROMERO",
    "CHRISTIAN ESTRADA", "MARTIN FELIX", "HIGINIO ESCALANTE",
}

def _hash(texto):
    return hashlib.md5(str(texto).encode()).hexdigest()

def _fecha_str(val):
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    if isinstance(val, str) and len(val) >= 8:
        return val[:10]
    return datetime.now().strftime("%Y-%m-%d")

def _monto(val):
    if val is None: return 0.0
    if isinstance(val, (int, float)): return float(val)
    try:
        s = str(val).replace("$","").replace(",","").strip()
        # Manejar formulas como =200+100
        if s.startswith("="):
            s = s[1:]
            # Evaluar suma simple
            if "+" in s:
                return sum(float(x.strip()) for x in s.split("+") if x.strip().replace(".","").isdigit())
            if s.replace(".","").isdigit():
                return float(s)
            return 0.0
        return float(s) if s else 0.0
    except: return 0.0

def _es_numero(val):
    if isinstance(val, (int, float)): return True
    if isinstance(val, str):
        s = str(val).replace("$","").replace(",","").strip()
        if s.startswith("="):
            return True
        try: float(s); return True
        except: return False
    return False

# TAB 1 — RELACION DE CARROS → registros + finanzas ingresos
def importar_relacion_carros(ws, conn):
    importados = duplicados = errores = 0
    NO_CLIENTES = {"S.SALUD","IMSS","CFE","GOBIERNO","MUNICIPIO",
                   "NONE","NOMBRE","CLIENTE","TOTAL","SUBTOTAL"}

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0] or not isinstance(row[0], (int, float)): continue
        try:
            folio    = int(row[0])
            fecha    = _fecha_str(row[1])
            nombre   = str(row[2] or "").strip().upper()
            auto     = str(row[3] or "").strip()
            placas   = str(row[4] or "").strip()
            servicio = str(row[5] or "").strip()
            refacc   = _monto(row[6])
            mano_ob  = _monto(row[7])
            cobrado  = _monto(row[8])

            if not nombre or len(nombre) < 3: continue
            if cobrado == 0 and refacc == 0 and mano_ob == 0: continue
            if nombre in NO_CLIENTES: continue
            if nombre.startswith("=") or nombre.startswith("TOTAL"): continue

            existe = conn.execute("SELECT id FROM registros WHERE numero_economico=?", (str(folio),)).fetchone()
            if existe:
                duplicados += 1
            else:
                conn.execute("""INSERT INTO registros
                    (nombre,telefono,correo,auto,numero_economico,placas,servicio,fecha,costo,estatus)
                    VALUES (?,?,?,?,?,?,?,?,?,?)""",
                    (nombre.title(),"","",auto,str(folio),placas,servicio,fecha,cobrado,"Entregado"))
                importados += 1

            # Ingreso en finanzas
            if cobrado > 0:
                try:
                    conn.execute("""INSERT INTO finanzas (tipo,concepto,monto,fecha,notas,hash_duplicado)
                        VALUES (?,?,?,?,?,?)""",
                        ("Ingreso", f"Folio {folio} - {nombre.title()} - {auto}",
                         cobrado, fecha, servicio, _hash(f"ing{folio}{fecha}{cobrado}")))
                except: pass

            # Refacciones como egreso
            if refacc > 0:
                try:
                    conn.execute("""INSERT INTO finanzas (tipo,concepto,monto,fecha,notas,hash_duplicado)
                        VALUES (?,?,?,?,?,?)""",
                        ("Egreso", f"Refacciones Folio {folio} - {auto}",
                         refacc, fecha, servicio, _hash(f"ref{folio}{fecha}{refacc}")))
                except: pass

        except Exception as ex:
            errores += 1
            error("ImportadorContador", f"Error fila relacion: {ex}")

    return importados, duplicados, errores


# TAB 2 — INFORME DIARIO → finanzas (ingresos izq, egresos der)
def importar_informe_diario(ws, conn):
    importados = duplicados = 0

    for row in ws.iter_rows(values_only=True):
        if not any(row): continue
        if row[0] == 'FECHA': continue
        if isinstance(row[0], str) and "SEMANA" in str(row[0]): continue

        # ── INGRESOS: fecha col0, nota col1, auto col2, cantidad col3
        try:
            fecha_val = row[0]
            monto_val = row[3]
            if isinstance(fecha_val, datetime) and _es_numero(monto_val):
                monto = _monto(monto_val)
                if monto > 0:
                    fecha    = _fecha_str(fecha_val)
                    nota     = str(row[1] or "")
                    auto     = str(row[2] or "")
                    concepto = f"Nota {nota} - {auto}".strip(" -") or "Ingreso del dia"
                    try:
                        conn.execute("""INSERT INTO finanzas (tipo,concepto,monto,fecha,notas,hash_duplicado)
                            VALUES (?,?,?,?,?,?)""",
                            ("Ingreso", concepto, monto, fecha, "Informe Diario",
                             _hash(f"ing_d{fecha}{nota}{monto}")))
                        importados += 1
                    except: duplicados += 1
        except: pass

        # ── EGRESOS: fecha col6, concepto col8, importe col11
        try:
            if len(row) > 11:
                fecha_val2 = row[6]
                monto_val2 = row[11]
                if isinstance(fecha_val2, datetime) and _es_numero(monto_val2):
                    monto = _monto(monto_val2)
                    if monto > 0:
                        fecha    = _fecha_str(fecha_val2)
                        concepto = str(row[8] or "Gasto operativo").strip()
                        if concepto and concepto != "CONCEPTO":
                            try:
                                conn.execute("""INSERT INTO finanzas (tipo,concepto,monto,fecha,notas,hash_duplicado)
                                    VALUES (?,?,?,?,?,?)""",
                                    ("Egreso", concepto, monto, fecha, "Informe Diario",
                                     _hash(f"eg_d{fecha}{concepto}{monto}")))
                                importados += 1
                            except: duplicados += 1
        except: pass

    return importados, duplicados


# TAB 3 — MANO DE OBRA → finanzas egresos
def importar_mano_obra(ws, conn):
    importados = duplicados = 0
    mecanico_actual = "Mecanico"
    fecha_actual = datetime.now().strftime("%Y-%m-%d")

    for row in ws.iter_rows(values_only=True):
        if not any(row): continue

        if isinstance(row[0], str) and "RELACION DE TRABAJOS" in row[0]:
            for emp in EMPLEADOS_CONOCIDOS:
                if emp in row[0].upper():
                    mecanico_actual = emp.title()
                    break
            continue

        if isinstance(row[0], datetime):
            fecha_actual = _fecha_str(row[0])

        trabajo = None
        monto   = 0.0

        # Trabajo en col2, monto en col4 o col5
        if len(row) > 2 and row[2] and isinstance(row[2], str) and len(row[2]) > 3:
            if "SUM" not in row[2].upper():
                trabajo = str(row[2]).strip()

        for col_idx in [4, 5]:
            if len(row) > col_idx and _es_numero(row[col_idx]):
                m = _monto(row[col_idx])
                if m > 0:
                    monto = m
                    break

        if trabajo and monto > 0:
            concepto = f"M.O. {mecanico_actual} - {trabajo}"
            try:
                conn.execute("""INSERT INTO finanzas (tipo,concepto,monto,fecha,notas,hash_duplicado)
                    VALUES (?,?,?,?,?,?)""",
                    ("Egreso", concepto, monto, fecha_actual, "Mano de Obra",
                     _hash(f"mo{fecha_actual}{mecanico_actual}{trabajo}{monto}")))
                importados += 1
            except: duplicados += 1

    return importados, duplicados


# TAB 4 — NOMINA → finanzas egresos + tabla nomina
def importar_nomina(ws, conn):
    importados = duplicados = 0
    semana_actual = ""
    fecha_actual  = datetime.now().strftime("%Y-%m-%d")

    conn.execute("""CREATE TABLE IF NOT EXISTS nomina (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        semana TEXT, nombre TEXT, sueldo REAL, bono REAL,
        prestamo REAL, adelanto REAL, faltas REAL, total REAL,
        fecha TEXT, hash_duplicado TEXT UNIQUE)""")

    for row in ws.iter_rows(values_only=True):
        if not any(row): continue
        if isinstance(row[0], str) and "NOMINA DE LA SEMANA" in row[0]:
            semana_actual = row[0].replace("NOMINA DE LA SEMANA DEL ", "").strip()
            continue
        if row[0] == 'NOMBRE': continue

        nombre = str(row[0] or "").strip()
        if not nombre or nombre in ("NOMBRE", " "): continue
        if not (nombre[0].isupper() and len(nombre) > 4): continue

        try:
            sueldo   = _monto(row[1]) if len(row) > 1 else 0
            bono     = _monto(row[2]) if len(row) > 2 else 0
            prestamo = _monto(row[3]) if len(row) > 3 else 0
            adelanto = _monto(row[4]) if len(row) > 4 else 0
            faltas   = _monto(row[5]) if len(row) > 5 else 0
            total    = sueldo + bono - prestamo - adelanto - faltas

            if sueldo == 0: continue

            hash_d = _hash(f"nom{semana_actual}{nombre}{sueldo}{bono}")
            try:
                conn.execute("""INSERT INTO nomina
                    (semana,nombre,sueldo,bono,prestamo,adelanto,faltas,total,fecha,hash_duplicado)
                    VALUES (?,?,?,?,?,?,?,?,?,?)""",
                    (semana_actual,nombre,sueldo,bono,prestamo,adelanto,faltas,total,fecha_actual,hash_d))
                importados += 1
            except: duplicados += 1

            if total > 0:
                try:
                    conn.execute("""INSERT INTO finanzas (tipo,concepto,monto,fecha,notas,hash_duplicado)
                        VALUES (?,?,?,?,?,?)""",
                        ("Egreso", f"Nomina - {nombre}", total, fecha_actual,
                         f"Semana {semana_actual}", _hash(f"nom_fin{semana_actual}{nombre}{total}")))
                except: pass

        except Exception as ex:
            error("ImportadorContador", f"Error nomina {nombre}: {ex}")

    return importados, duplicados


# FUNCION PRINCIPAL
def importar_excel_contador(ruta_archivo):
    if not os.path.exists(ruta_archivo):
        error("ImportadorContador", f"Archivo no encontrado: {ruta_archivo}")
        return None
    try:
        wb   = openpyxl.load_workbook(ruta_archivo, data_only=True)
        conn = conectar()
        conn.execute("PRAGMA journal_mode=WAL")
        resultados = {}

        if 'RELACION DE CARROS' in wb.sheetnames:
            imp, dup, err = importar_relacion_carros(wb['RELACION DE CARROS'], conn)
            resultados['relacion_carros'] = {"importados": imp, "duplicados": dup, "errores": err}
            info("ImportadorContador", f"Relacion Carros: {imp} importados, {dup} duplicados")

        if 'INFORME DIARIO' in wb.sheetnames:
            imp, dup = importar_informe_diario(wb['INFORME DIARIO'], conn)
            resultados['informe_diario'] = {"importados": imp, "duplicados": dup}
            info("ImportadorContador", f"Informe Diario: {imp} importados, {dup} duplicados")

        if 'MANO DE OBRA' in wb.sheetnames:
            imp, dup = importar_mano_obra(wb['MANO DE OBRA'], conn)
            resultados['mano_obra'] = {"importados": imp, "duplicados": dup}
            info("ImportadorContador", f"Mano de Obra: {imp} importados, {dup} duplicados")

        if 'NOMINA' in wb.sheetnames:
            imp, dup = importar_nomina(wb['NOMINA'], conn)
            resultados['nomina'] = {"importados": imp, "duplicados": dup}
            info("ImportadorContador", f"Nomina: {imp} importados, {dup} duplicados")

        conn.commit()
        conn.close()
        info("ImportadorContador", f"Importacion completa: {ruta_archivo}")
        return resultados

    except Exception as ex:
        error("ImportadorContador", f"Error general: {ex}")
        return None
