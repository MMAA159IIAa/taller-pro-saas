import openpyxl
import hashlib
import sqlite3
import os
import sys
from datetime import datetime

# ── Path setup ────────────────────────────────────────────────────────────────
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from utils.database import conectar
from utils.logger import info, error, warn

EMPLEADOS_CONOCIDOS = {
    "CARLOS DANIEL GUILLEN MONTAÑO", "FIDEL ENRIQUE LARA LOPEZ",
    "GUADALUPE LOREDO BELTRAN", "JESUS ENRIQUE RAMIREZ ROMERO",
    "CHRISTIAN ESTRADA", "MARTIN FELIX", "HIGINIO ESCALANTE",
}

def _hash(texto):
    return hashlib.md5(str(texto).encode()).hexdigest()

def _fecha_str(val):
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    if isinstance(val, str) and len(val) >= 8:
        return val[:10]
    return datetime.now().strftime("%Y-%m-%d")

def _monto(val):
    if val is None: return 0.0
    if isinstance(val, (int, float)): return float(val)
    try: return float(str(val).replace("$","").replace(",","").replace("=","").strip())
    except: return 0.0

# ══════════════════════════════════════════════════════════════════════════════
# TAB 1 — RELACION DE CARROS → tabla registros
# ══════════════════════════════════════════════════════════════════════════════
def importar_relacion_carros(ws, conn):
    importados = duplicados = errores = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0] or not isinstance(row[0], (int, float)): continue
        try:
            folio    = int(row[0])
            fecha    = _fecha_str(row[1])
            nombre   = str(row[2] or "").strip()
            auto     = str(row[3] or "").strip()
            placas   = str(row[4] or "").strip()
            servicio = str(row[5] or "").strip()
            refacc   = _monto(row[6])
            mano_ob  = _monto(row[7])
            cobrado  = _monto(row[8])

            if not nombre or cobrado == 0: continue

            hash_dup = _hash(f"reg{folio}{fecha}{nombre}")
            try:
                conn.execute("""
                    INSERT INTO registros
                    (nombre, telefono, correo, auto, numero_economico, placas,
                     servicio, fecha, costo, estatus)
                    VALUES (?,?,?,?,?,?,?,?,?,?)
                """, (nombre, "", "", auto, str(folio), placas,
                      servicio, fecha, cobrado, "Entregado"))
                importados += 1
            except sqlite3.IntegrityError:
                duplicados += 1
            except Exception as ex:
                # Si no hay UNIQUE constraint en registros, verificar manualmente
                existe = conn.execute("""
                    SELECT id FROM registros
                    WHERE numero_economico=? AND fecha=? AND nombre=?
                """, (str(folio), fecha, nombre)).fetchone()
                if not existe:
                    conn.execute("""
                        INSERT INTO registros
                        (nombre, telefono, correo, auto, numero_economico, placas,
                         servicio, fecha, costo, estatus)
                        VALUES (?,?,?,?,?,?,?,?,?,?)
                    """, (nombre, "", "", auto, str(folio), placas,
                          servicio, fecha, cobrado, "Entregado"))
                    importados += 1
                else:
                    duplicados += 1

            # También registrar en finanzas como ingreso
            hash_fin = _hash(f"ing{folio}{fecha}{cobrado}")
            try:
                conn.execute("""
                    INSERT INTO finanzas (tipo, concepto, monto, fecha, notas, hash_duplicado)
                    VALUES (?,?,?,?,?,?)
                """, ("Ingreso", f"Folio {folio} - {nombre} - {auto}",
                      cobrado, fecha, servicio, hash_fin))
            except: pass

            # Refacciones como egreso
            if refacc > 0:
                hash_ref = _hash(f"ref{folio}{fecha}{refacc}")
                try:
                    conn.execute("""
                        INSERT INTO finanzas (tipo, concepto, monto, fecha, notas, hash_duplicado)
                        VALUES (?,?,?,?,?,?)
                    """, ("Egreso", f"Refacciones - Folio {folio} - {auto}",
                          refacc, fecha, f"Refacciones {servicio}", hash_ref))
                except: pass

        except Exception as ex:
            errores += 1
            error("ImportadorContador", f"Error en fila relacion carros: {ex}")

    return importados, duplicados, errores


# ══════════════════════════════════════════════════════════════════════════════
# TAB 2 — INFORME DIARIO → tabla finanzas
# Izquierda = ingresos (col D), Derecha = egresos (col L)
# ══════════════════════════════════════════════════════════════════════════════
def importar_informe_diario(ws, conn):
    importados = duplicados = 0
    fecha_actual = datetime.now().strftime("%Y-%m-%d")

    for row in ws.iter_rows(values_only=True):
        # Detectar encabezado de semana para contexto
        if row[0] == ' ' and row[3] and "SEMANA" in str(row[3]):
            continue
        if row[0] == 'FECHA': continue
        if not any(row): continue

        # INGRESOS — columnas: fecha(0), nota(1), auto(2), cantidad(3)
        try:
            if isinstance(row[0], datetime) and isinstance(row[3], (int, float)) and row[3] > 0:
                fecha  = _fecha_str(row[0])
                nota   = str(row[1] or "")
                auto   = str(row[2] or "")
                monto  = _monto(row[3])
                concepto = f"Nota {nota} - {auto}".strip(" -")
                hash_d = _hash(f"diario_ing{fecha}{nota}{monto}")
                try:
                    conn.execute("""
                        INSERT INTO finanzas (tipo,concepto,monto,fecha,notas,hash_duplicado)
                        VALUES (?,?,?,?,?,?)
                    """, ("Ingreso", concepto, monto, fecha, "Informe Diario", hash_d))
                    importados += 1
                except: duplicados += 1
        except: pass

        # EGRESOS — columnas: fecha(6), concepto(8), importe(11)
        try:
            if isinstance(row[6], datetime) and isinstance(row[11], (int, float)) and row[11] > 0:
                fecha    = _fecha_str(row[6])
                concepto = str(row[8] or "Gasto operativo").strip()
                monto    = _monto(row[11])
                hash_d   = _hash(f"diario_eg{fecha}{concepto}{monto}")
                try:
                    conn.execute("""
                        INSERT INTO finanzas (tipo,concepto,monto,fecha,notas,hash_duplicado)
                        VALUES (?,?,?,?,?,?)
                    """, ("Egreso", concepto, monto, fecha, "Informe Diario", hash_d))
                    importados += 1
                except: duplicados += 1
        except: pass

    return importados, duplicados


# ══════════════════════════════════════════════════════════════════════════════
# TAB 3 — MANO DE OBRA → tabla finanzas (egresos mano de obra)
# ══════════════════════════════════════════════════════════════════════════════
def importar_mano_obra(ws, conn):
    importados = duplicados = 0
    mecanico_actual = "Mecanico"
    fecha_actual    = datetime.now().strftime("%Y-%m-%d")

    for row in ws.iter_rows(values_only=True):
        if not any(row): continue

        # Detectar nombre del mecánico en encabezado
        if isinstance(row[0], str) and "RELACION DE TRABAJOS" in row[0]:
            for emp in EMPLEADOS_CONOCIDOS:
                if emp in row[0].upper():
                    mecanico_actual = emp.title()
                    break
            continue

        # Detectar fecha
        if isinstance(row[0], datetime):
            fecha_actual = _fecha_str(row[0])

        # Detectar trabajo y monto (col 2 y col 4 o 5)
        trabajo = None
        monto   = 0.0
        if row[2] and isinstance(row[2], str) and len(row[2]) > 3:
            trabajo = str(row[2]).strip()
        for col_idx in [4, 5]:
            if len(row) > col_idx and isinstance(row[col_idx], (int, float)) and row[col_idx] > 0:
                monto = float(row[col_idx])
                break

        if trabajo and monto > 0 and "SUM" not in str(monto):
            concepto = f"M.O. {mecanico_actual} - {trabajo}"
            hash_d   = _hash(f"mo{fecha_actual}{mecanico_actual}{trabajo}{monto}")
            try:
                conn.execute("""
                    INSERT INTO finanzas (tipo,concepto,monto,fecha,notas,hash_duplicado)
                    VALUES (?,?,?,?,?,?)
                """, ("Egreso", concepto, monto, fecha_actual, "Mano de Obra", hash_d))
                importados += 1
            except:
                duplicados += 1

    return importados, duplicados


# ══════════════════════════════════════════════════════════════════════════════
# TAB 4 — NOMINA → tabla finanzas (egresos nomina) + tabla empleados
# ══════════════════════════════════════════════════════════════════════════════
def importar_nomina(ws, conn):
    importados = duplicados = 0
    semana_actual = ""
    fecha_actual  = datetime.now().strftime("%Y-%m-%d")

    # Crear tabla empleados si no existe
    conn.execute("""
        CREATE TABLE IF NOT EXISTS nomina (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            semana TEXT, nombre TEXT, sueldo REAL,
            bono REAL, prestamo REAL, adelanto REAL,
            faltas REAL, total REAL, fecha TEXT,
            hash_duplicado TEXT UNIQUE)
    """)

    for row in ws.iter_rows(values_only=True):
        if not any(row): continue

        # Detectar encabezado de semana
        if isinstance(row[0], str) and "NOMINA DE LA SEMANA" in row[0]:
            semana_actual = row[0].replace("NOMINA DE LA SEMANA DEL ", "").strip()
            # Extraer fecha aproximada de la semana
            continue

        if row[0] == 'NOMBRE': continue

        # Detectar empleado
        nombre = str(row[0] or "").strip()
        if not nombre or nombre in ("NOMBRE", " "): continue
        if not any(emp in nombre.upper() for emp in [e.split()[0] for e in EMPLEADOS_CONOCIDOS]):
            # Verificar si parece nombre de persona
            if not (nombre[0].isupper() and len(nombre) > 5): continue

        try:
            sueldo   = _monto(row[1])
            bono     = _monto(row[2])
            prestamo = _monto(row[3])
            adelanto = _monto(row[4])
            faltas   = _monto(row[5])
            total    = sueldo + bono - prestamo - adelanto - faltas

            if sueldo == 0: continue

            hash_d = _hash(f"nom{semana_actual}{nombre}{sueldo}{bono}")
            try:
                conn.execute("""
                    INSERT INTO nomina
                    (semana, nombre, sueldo, bono, prestamo, adelanto, faltas, total, fecha, hash_duplicado)
                    VALUES (?,?,?,?,?,?,?,?,?,?)
                """, (semana_actual, nombre, sueldo, bono, prestamo,
                      adelanto, faltas, total, fecha_actual, hash_d))
                importados += 1
            except:
                duplicados += 1

            # Registrar en finanzas como egreso
            if total > 0:
                hash_fin = _hash(f"nom_fin{semana_actual}{nombre}{total}")
                try:
                    conn.execute("""
                        INSERT INTO finanzas (tipo,concepto,monto,fecha,notas,hash_duplicado)
                        VALUES (?,?,?,?,?,?)
                    """, ("Egreso", f"Nomina - {nombre}",
                          total, fecha_actual, f"Semana {semana_actual}", hash_fin))
                except: pass

        except Exception as ex:
            error("ImportadorContador", f"Error nomina {nombre}: {ex}")

    return importados, duplicados


# ══════════════════════════════════════════════════════════════════════════════
# FUNCIÓN PRINCIPAL
# ══════════════════════════════════════════════════════════════════════════════
def importar_excel_contador(ruta_archivo):
    """
    Importa el Excel completo del contador (4 tabs).
    Retorna dict con resumen de resultados.
    """
    if not os.path.exists(ruta_archivo):
        error("ImportadorContador", f"Archivo no encontrado: {ruta_archivo}")
        return None

    try:
        wb   = openpyxl.load_workbook(ruta_archivo, data_only=True)
        conn = conectar()
        conn.execute("PRAGMA journal_mode=WAL")
        resultados = {}

        # TAB 1 - Relacion de Carros
        if 'RELACION DE CARROS' in wb.sheetnames:
            imp, dup, err = importar_relacion_carros(wb['RELACION DE CARROS'], conn)
            resultados['relacion_carros'] = {"importados": imp, "duplicados": dup, "errores": err}
            info("ImportadorContador", f"Relacion Carros: {imp} importados, {dup} duplicados")

        # TAB 2 - Informe Diario
        if 'INFORME DIARIO' in wb.sheetnames:
            imp, dup = importar_informe_diario(wb['INFORME DIARIO'], conn)
            resultados['informe_diario'] = {"importados": imp, "duplicados": dup}
            info("ImportadorContador", f"Informe Diario: {imp} importados, {dup} duplicados")

        # TAB 3 - Mano de Obra
        if 'MANO DE OBRA' in wb.sheetnames:
            imp, dup = importar_mano_obra(wb['MANO DE OBRA'], conn)
            resultados['mano_obra'] = {"importados": imp, "duplicados": dup}
            info("ImportadorContador", f"Mano de Obra: {imp} importados, {dup} duplicados")

        # TAB 4 - Nomina
        if 'NOMINA' in wb.sheetnames:
            imp, dup = importar_nomina(wb['NOMINA'], conn)
            resultados['nomina'] = {"importados": imp, "duplicados": dup}
            info("ImportadorContador", f"Nomina: {imp} importados, {dup} duplicados")

        conn.commit()
        conn.close()
        info("ImportadorContador", f"Importacion completa: {ruta_archivo}")
        return resultados

    except Exception as ex:
        error("ImportadorContador", f"Error general: {ex}")
        return None
