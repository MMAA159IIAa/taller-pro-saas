import openpyxl
import sqlite3
import os
import sys
from datetime import datetime

# Setup paths to import the agent
sys.path.insert(0, os.getcwd())
try:
    from agentes.contador_import_agent import importar_excel_contador
except ImportError as e:
    print(f"Error importando agente: {e}")
    sys.exit(1)

# Usar una base de datos temporal para pruebas
db_test = "test_import.db"
if os.path.exists(db_test): os.remove(db_test)

def crear_tablas_test(conn):
    c = conn.cursor()
    c.execute("""CREATE TABLE IF NOT EXISTS registros (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        nombre TEXT, telefono TEXT, correo TEXT,
        auto TEXT, numero_economico TEXT, placas TEXT,
        servicio TEXT, fecha TEXT, costo REAL DEFAULT 0,
        estatus TEXT DEFAULT 'Pendiente')""")
    c.execute("""CREATE TABLE IF NOT EXISTS finanzas (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        tipo TEXT, concepto TEXT, monto REAL,
        fecha TEXT, notas TEXT,
        hash_duplicado TEXT UNIQUE)""")
    conn.commit()

# Parcheamos la función conectar del agente para que use nuestra DB de test
import agentes.contador_import_agent
agentes.contador_import_agent.conectar = lambda: sqlite3.connect(db_test)

# También parcheamos el logger para ver errores en consola
def log_error(mod, msg): print(f"[DB_ERROR] {mod}: {msg}")
agentes.contador_import_agent.error = log_error

ruta_excel = r"c:\Users\leech\OneDrive\Documentos\RELACION DE CARROS 2026.xlsx"
print(f"Probando importación de: {ruta_excel}")

res = importar_excel_contador(ruta_excel)

print(f"\nResultados del importador: {res}")

if res:
    conn = sqlite3.connect(db_test)
    print("\n--- Conteo de Registros en DB ---")
    reg_count = conn.execute("SELECT COUNT(*) FROM registros").fetchone()[0]
    fin_ing_count = conn.execute("SELECT COUNT(*) FROM finanzas WHERE tipo='Ingreso'").fetchone()[0]
    fin_eg_count = conn.execute("SELECT COUNT(*) FROM finanzas WHERE tipo='Egreso'").fetchone()[0]
    print(f"Clientes (registros): {reg_count}")
    print(f"Ingresos (finanzas): {fin_ing_count}")
    print(f"Egresos (finanzas): {fin_eg_count}")
    
    if fin_eg_count == 0:
        print("\nADVERTENCIA: No se detectaron egresos (gastos).")
        # Vamos a ver por qué falló el Informe Diario
        wb = openpyxl.load_workbook(ruta_excel, data_only=True)
        if 'INFORME DIARIO' in wb.sheetnames:
            ws = wb['INFORME DIARIO']
            print("\nInspeccionando INFORME DIARIO...")
            rows = list(ws.iter_rows(values_only=True))
            for i, row in enumerate(rows[:10]):
                print(f"Fila {i}: {row}")
    conn.close()
else:
    print("La función importar_excel_contador devolvió None")
