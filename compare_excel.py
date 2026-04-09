import openpyxl
import os

files = [
    r"c:\Users\leech\OneDrive\Documentos\RELACION DE CARROS 2025.xlsx",
    r"c:\Users\leech\OneDrive\Documentos\RELACION DE CARROS 2026.xlsx"
]

for ruta in files:
    if os.path.exists(ruta):
        print(f"\n=== Archivo: {os.path.basename(ruta)} ===")
        wb = openpyxl.load_workbook(ruta, read_only=True, data_only=True)
        print(f"Hojas: {wb.sheetnames}")
        if 'INFORME DIARIO' in wb.sheetnames:
            ws = wb['INFORME DIARIO']
            print("--- Primeras 3 filas de INFORME DIARIO ---")
            for row in ws.iter_rows(max_row=3, values_only=True):
                print(row)
    else:
        print(f"No encontrado: {ruta}")
