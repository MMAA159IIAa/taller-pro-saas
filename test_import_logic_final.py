import openpyxl
import os
import sqlite3
from datetime import datetime

base = r"c:\Users\leech\OneDrive\Documentos\taller_pro_v3"
db_path = os.path.join(base, "test_import_v3.db")
if os.path.exists(db_path): os.remove(db_path)

conn = sqlite3.connect(db_path)
conn.execute("CREATE TABLE registros (id INTEGER PRIMARY KEY, nombre TEXT, auto TEXT, numero_economico TEXT, servicio TEXT, fecha TEXT, costo REAL, placas TEXT)")
conn.execute("CREATE TABLE finanzas (id INTEGER PRIMARY KEY, tipo TEXT, concepto TEXT, monto REAL, fecha TEXT, Notas TEXT)")

def _monto(val):
    if val is None: return 0.0
    if isinstance(val, (int, float)): return float(val)
    try: return float(str(val).replace("$","").replace(",","").replace("=","").strip())
    except: return 0.0

def _fecha_str(val):
    if isinstance(val, datetime): return val.strftime("%Y-%m-%d")
    return str(val)[:10]

def test_logic():
    xlsx = r"c:\Users\leech\OneDrive\Documentos\RELACION DE CARROS 2026.xlsx"
    wb = openpyxl.load_workbook(xlsx, data_only=True)
    # Buscamos la hoja correcta
    ws = None
    for name in wb.sheetnames:
        if "INFORME" in name.upper() or "RELACION" in name.upper():
            ws = wb[name]
            break
    
    if not ws:
        print("Hoja no encontrada")
        return

    print(f"Probando con hoja: {ws.title}")
    
    for row in ws.iter_rows(min_row=1, max_row=50, values_only=True):
        if not any(row): continue
        
        # FLUJO IZQUIERDO (Indices 0, 1, 2, 3)
        # 0: Fecha, 1: Folio, 2: Carro, 3: Ingreso
        if (isinstance(row[0], datetime) or (row[0] and isinstance(row[0], (int, float)))) and len(str(row[2] or "")) > 2:
            nombre = str(row[2]).strip()
            if nombre.upper() not in ["CONCEPTO", "NOMBRE", "FECHA", "CANTIDAD", "NOTA DE VENTA"]:
                fecha = _fecha_str(row[0])
                monto = _monto(row[3])
                print(f"[CLIENTE] {fecha} | {nombre} | ${monto}")
                conn.execute("INSERT INTO registros (nombre, auto, servicio, fecha, costo) VALUES (?,?,?,?,?)", (nombre, nombre, "Importado", fecha, monto))

        # FLUJO DERECHO (Indices 6, 8, 11)
        # 6: Fecha, 8: Concepto, 11: Importe
        if len(row) > 11 and isinstance(row[6], datetime) and row[8] and _monto(row[11]) > 0:
            concepto = str(row[8]).strip()
            if concepto.upper() not in ["CONCEPTO", "NOMBRE", "FECHA", "CANTIDAD", "IMPORTE"]:
                fecha = _fecha_str(row[6])
                monto = _monto(row[11])
                print(f"  [FINANZA] {fecha} | {concepto} | ${monto}")
                conn.execute("INSERT INTO finanzas (tipo, concepto, monto, fecha) VALUES (?,?,?,?)", ("Egreso", concepto, monto, fecha))

    conn.commit()
    conn.close()

if __name__ == "__main__":
    test_logic()
